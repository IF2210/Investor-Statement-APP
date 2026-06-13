"""
StatementIQ US — PE/VC LP Statement Generator
Vercel Serverless Function (Python)

POST /api/generate
Accepts LP investor JSON, returns Excel file + sends email copy.
Supports multi-currency: USD, EUR, GBP, CAD, AUD, JPY, CHF, SGD
"""

from http.server import BaseHTTPRequestHandler
import json, io, os, smtplib, base64
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Palette ────────────────────────────────────────────────────
NAVY       = "0D1B2A"
MID_NAVY   = "1B3A5C"
GOLD       = "C9A84C"
GOLD_LIGHT = "F0DFA0"
BLUE_LIGHT = "D6E4F0"
GREY_LIGHT = "F5F7FA"
GREY_MID   = "E8EBF0"
WHITE      = "FFFFFF"
GREEN      = "1A4731"
GREEN_BG   = "C6F6D5"
RED        = "9B2335"
RED_BG     = "FED7D7"
TEXT       = "1A202C"
TEXT_MED   = "4A5568"

# ── Currency symbols & formats ─────────────────────────────────
CURRENCIES = {
    "USD": {"symbol": "$",  "code": "USD", "format": '$#,##0.00',    "int": '$#,##0'},
    "EUR": {"symbol": "€",  "code": "EUR", "format": '€#,##0.00',    "int": '€#,##0'},
    "GBP": {"symbol": "£",  "code": "GBP", "format": '£#,##0.00',    "int": '£#,##0'},
    "CAD": {"symbol": "C$", "code": "CAD", "format": 'C$#,##0.00',   "int": 'C$#,##0'},
    "AUD": {"symbol": "A$", "code": "AUD", "format": 'A$#,##0.00',   "int": 'A$#,##0'},
    "JPY": {"symbol": "¥",  "code": "JPY", "format": '¥#,##0',       "int": '¥#,##0'},
    "CHF": {"symbol": "Fr", "code": "CHF", "format": 'Fr#,##0.00',   "int": 'Fr#,##0'},
    "SGD": {"symbol": "S$", "code": "SGD", "format": 'S$#,##0.00',   "int": 'S$#,##0'},
}

def get_curr(code):
    return CURRENCIES.get(code.upper(), CURRENCIES["USD"])

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def side(color="CBD5E0"):
    return Side(style="thin", color=color)

def cell(ws, row, col, value=None, bold=False, italic=False,
         size=10, color=TEXT, bg=None, align="left", valign="center",
         number_format=None, border=None, wrap=False, indent=0):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical=valign,
                             wrap_text=wrap, indent=indent)
    if bg:   c.fill = solid(bg)
    if number_format: c.number_format = number_format
    if border: c.border = border
    return c

def hdr_border():
    s = side()
    return Border(bottom=s)

def full_border():
    s = side()
    return Border(top=s, bottom=s, left=s, right=s)


def generate_excel(investor: dict) -> bytes:
    curr     = get_curr(investor.get("currency", "USD"))
    sym      = curr["symbol"]
    fmt_dec  = curr["format"]
    fmt_int  = curr["int"]
    pct_fmt  = "0.00%"

    wb = Workbook()
    ws = wb.active
    ws.title = "LP Statement"
    ws.sheet_view.showGridLines = False

    # Column widths
    for i, w in enumerate([2, 26, 16, 16, 16, 16, 14, 2], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, 90):
        ws.row_dimensions[r].height = 15

    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 52
    ws.row_dimensions[3].height = 5

    # ── HEADER BANNER ──────────────────────────────────────────
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.fill = solid(NAVY)
    c.value = "LIMITED PARTNER CAPITAL ACCOUNT STATEMENT"
    c.font  = Font(name="Calibri", bold=True, size=17, color=GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for col in range(2, 8):
        ws.cell(row=3, column=col).fill = solid(GOLD)

    # ── LP META ────────────────────────────────────────────────
    fund_name  = investor.get("fund_name", "")
    lp_name    = investor.get("lp_name", "")
    acct       = investor.get("account_number", "")
    email      = investor.get("email", "")
    period_f   = investor.get("period_from", "")
    period_t   = investor.get("period_to", "")
    vintage    = investor.get("vintage_year", "")
    commitment = float(investor.get("commitment", 0))

    meta = [
        ("Fund Name",         fund_name),
        ("Limited Partner",   lp_name),
        ("Account Number",    acct),
        ("Currency",          curr["code"]),
        ("Statement Period",  f"{period_f}  –  {period_t}"),
        ("Vintage Year",      str(vintage) if vintage else "—"),
        ("Total Commitment",  None),   # will be formula
        ("Report Date",       datetime.today().strftime("%B %d, %Y")),
    ]

    b = Border(bottom=side("D0D0D0"))
    for i, (lbl, val) in enumerate(meta):
        row = 5 + i
        ws.row_dimensions[row].height = 18
        cell(ws, row, 2, lbl, bold=True, color=TEXT_MED, size=9, bg=GREY_LIGHT, border=b)
        ws.merge_cells(f"C{row}:G{row}")
        if lbl == "Total Commitment":
            c = ws.cell(row=row, column=3)
            c.value = commitment
            c.font  = Font(name="Calibri", bold=True, size=10, color=TEXT)
            c.number_format = fmt_int
            c.fill = solid(GREY_LIGHT)
            c.border = b
        else:
            cell(ws, row, 3, val, size=10, color=TEXT, bg=GREY_LIGHT, border=b)

    ws.row_dimensions[14].height = 10

    # ── SUMMARY KPI CARDS (row 15-17) ─────────────────────────
    ws.row_dimensions[15].height = 20
    ws.merge_cells("B15:G15")
    cell(ws, 15, 2, "CAPITAL ACCOUNT SUMMARY", bold=True, size=11,
         color=WHITE, bg=MID_NAVY, align="center")

    ws.row_dimensions[16].height = 18
    ws.row_dimensions[17].height = 28

    card_defs = [
        ("Opening Value",   BLUE_LIGHT, "B"),
        ("Closing Value",   GOLD_LIGHT, "D"),
        ("Net Gain / (Loss)", GREY_LIGHT, "F"),
        ("Return",          GREEN_BG,   "G"),
    ]
    for label, bg, col in card_defs:
        end_col = chr(ord(col) + 1) if col != "G" else "G"
        if col != "G":
            ws.merge_cells(f"{col}16:{end_col}16")
            ws.merge_cells(f"{col}17:{end_col}17")
        cell(ws, 16, ord(col)-64, label, bold=True, size=9,
             color=TEXT_MED, bg=bg, align="center")

    ws.row_dimensions[18].height = 8

    # ── CAPITAL ACCOUNT TABLE (row 19+) ───────────────────────
    ws.row_dimensions[19].height = 20
    ws.merge_cells("B19:G19")
    cell(ws, 19, 2, "CAPITAL ACCOUNT ACTIVITY", bold=True, size=11,
         color=WHITE, bg=MID_NAVY, align="center")

    ws.row_dimensions[20].height = 18
    col_heads = ["", "Description", "Date", "Shares / Units",
                 f"Price / NAV ({sym})", f"Amount ({sym})", ""]
    for col, h in enumerate(col_heads, 1):
        cell(ws, 20, col, h, bold=True, size=9, color=WHITE, bg=NAVY,
             align="center" if col > 2 else "left")

    # Opening balance row
    ws.row_dimensions[21].height = 18
    opening_shares = float(investor.get("opening_shares", 0))
    opening_nav    = float(investor.get("opening_nav", 0))
    cell(ws, 21, 2, "Opening Balance", bold=True, size=10, bg=GREY_LIGHT)
    cell(ws, 21, 3, period_f, size=9, color=TEXT_MED, bg=GREY_LIGHT, align="center")
    cell(ws, 21, 4, opening_shares, size=10, bg=GREY_LIGHT, align="right",
         number_format="#,##0.0000")
    cell(ws, 21, 5, opening_nav, size=10, bg=GREY_LIGHT, align="right",
         number_format=fmt_dec)
    cell(ws, 21, 6, "=D21*E21", bold=True, size=10, bg=BLUE_LIGHT,
         align="right", number_format=fmt_int)

    ws.row_dimensions[22].height = 8

    # ── CAPITAL ACTIVITY ───────────────────────────────────────
    ws.row_dimensions[23].height = 20
    ws.merge_cells("B23:G23")
    cell(ws, 23, 2, "CAPITAL CALLS & DISTRIBUTIONS", bold=True, size=11,
         color=WHITE, bg=MID_NAVY, align="center")

    ws.row_dimensions[24].height = 18
    txn_heads = ["", "Transaction Type", "Date", "Shares / Units",
                 f"Price ({sym})", f"Amount ({sym})", ""]
    for col, h in enumerate(txn_heads, 1):
        cell(ws, 24, col, h, bold=True, size=9, color=WHITE, bg=NAVY,
             align="center" if col > 2 else "left")

    transactions = investor.get("transactions", [])
    row = 25
    txn_rows = []

    if not transactions:
        ws.merge_cells(f"B{row}:G{row}")
        ws.row_dimensions[row].height = 18
        cell(ws, row, 2, "No capital activity during this period.",
             italic=True, color=TEXT_MED, size=9, align="center")
        row += 1
    else:
        for idx, txn in enumerate(transactions):
            bg_row   = WHITE if idx % 2 == 0 else GREY_LIGHT
            is_call  = str(txn.get("type", "")).lower() in ["capital call", "call", "contribution"]
            units    = float(txn.get("shares", txn.get("units", 0)))
            price    = float(txn.get("price", txn.get("nav", 0)))
            amount   = units * price if is_call else -(units * price)
            t_color  = GREEN if is_call else RED
            t_bg     = GREEN_BG if is_call else RED_BG
            label    = "Capital Call" if is_call else "Distribution"

            ws.row_dimensions[row].height = 18
            cell(ws, row, 2, label, bold=True, size=10, color=t_color, bg=bg_row)
            cell(ws, row, 3, txn.get("date", ""), size=9, color=TEXT_MED,
                 bg=bg_row, align="center")
            cell(ws, row, 4, units, size=10, bg=bg_row, align="right",
                 number_format="#,##0.0000")
            cell(ws, row, 5, price, size=10, bg=bg_row, align="right",
                 number_format=fmt_dec)
            cell(ws, row, 6, amount, bold=True, size=10, color=t_color,
                 bg=t_bg, align="right",
                 number_format=f'{sym}#,##0;({sym}#,##0);-')
            txn_rows.append(row)
            row += 1

    # Net activity row
    ws.row_dimensions[row].height = 18
    cell(ws, row, 2, "Net Capital Activity", bold=True, size=10, color=WHITE, bg=MID_NAVY)
    if txn_rows:
        cell(ws, row, 4,
             f"=SUM(D25:D{row-1})", bold=True, size=10, color=WHITE,
             bg=MID_NAVY, align="right", number_format="#,##0.0000")
        cell(ws, row, 6,
             f"=SUM(F25:F{row-1})", bold=True, size=10, color=WHITE,
             bg=MID_NAVY, align="right",
             number_format=f'{sym}#,##0;({sym}#,##0);-')
    net_row = row
    row += 1
    ws.row_dimensions[row].height = 8

    # ── CLOSING POSITION ───────────────────────────────────────
    close_hdr = row + 1
    ws.row_dimensions[close_hdr].height = 20
    ws.merge_cells(f"B{close_hdr}:G{close_hdr}")
    cell(ws, close_hdr, 2, "CLOSING POSITION", bold=True, size=11,
         color=WHITE, bg=MID_NAVY, align="center")

    chdr = close_hdr + 1
    ws.row_dimensions[chdr].height = 18
    for col, h in enumerate(["", "Description", "Date", "Shares / Units",
                              f"Closing NAV ({sym})", f"Closing Value ({sym})", ""], 1):
        cell(ws, chdr, col, h, bold=True, size=9, color=WHITE, bg=NAVY,
             align="center" if col > 2 else "left")

    cval = chdr + 1
    ws.row_dimensions[cval].height = 24
    closing_nav = float(investor.get("closing_nav", 0))

    txn_shares = sum(
        float(t.get("shares", t.get("units", 0)))
        if str(t.get("type","")).lower() in ["capital call","call","contribution"]
        else -float(t.get("shares", t.get("units", 0)))
        for t in transactions
    )
    closing_shares = opening_shares + txn_shares

    cell(ws, cval, 2, "Closing Balance", bold=True, size=11, color=NAVY, bg=BLUE_LIGHT)
    cell(ws, cval, 3, period_t, size=9, color=TEXT_MED, bg=BLUE_LIGHT, align="center")
    cell(ws, cval, 4, f"=D21+D{net_row}", bold=True, size=11, color=NAVY,
         bg=BLUE_LIGHT, align="right", number_format="#,##0.0000")
    cell(ws, cval, 5, closing_nav, bold=True, size=11, color=NAVY,
         bg=BLUE_LIGHT, align="right", number_format=fmt_dec)
    cell(ws, cval, 6, f"=D{cval}*E{cval}", bold=True, size=13, color=NAVY,
         bg=GOLD_LIGHT, align="right", number_format=fmt_int)

    row = cval + 1
    ws.row_dimensions[row].height = 8

    # ── PERFORMANCE SUMMARY ────────────────────────────────────
    perf_hdr = row + 1
    ws.row_dimensions[perf_hdr].height = 20
    ws.merge_cells(f"B{perf_hdr}:G{perf_hdr}")
    cell(ws, perf_hdr, 2, "PERFORMANCE SUMMARY", bold=True, size=11,
         color=WHITE, bg=MID_NAVY, align="center")

    perf_items = [
        ("Opening Value",              f"=F21",                                       fmt_int,  False),
        ("Total Capital Called",       f"=SUMIF(F25:F{net_row-1},\">0\",F25:F{net_row-1})", fmt_int, False),
        ("Total Distributions",        f"=SUMIF(F25:F{net_row-1},\"<0\",F25:F{net_row-1})", f'{sym}#,##0;({sym}#,##0);-', False),
        ("Closing Value",              f"=F{cval}",                                   fmt_int,  False),
        ("Net Gain / (Loss)",          f"=F{cval}-F{perf_hdr+1}-F{perf_hdr+2}",      f'{sym}#,##0;({sym}#,##0);-', True),
        ("Return on Opening Value",    f"=IF(F{perf_hdr+1}=0,0,(F{cval}-F{perf_hdr+1}-F{net_row})/F{perf_hdr+1})", pct_fmt, True),
    ]

    pr = perf_hdr + 1
    for i, (lbl, frm, fmt, highlight) in enumerate(perf_items):
        bg = GREY_LIGHT if i % 2 == 0 else WHITE
        ws.row_dimensions[pr].height = 18
        ws.merge_cells(f"B{pr}:E{pr}")
        cell(ws, pr, 2, lbl, bold=highlight, size=10,
             color=NAVY if highlight else TEXT, bg=GOLD_LIGHT if highlight else bg)
        cell(ws, pr, 6, frm, bold=highlight, size=11 if highlight else 10,
             color=NAVY, bg=GOLD_LIGHT if highlight else bg,
             align="right", number_format=fmt)
        pr += 1

    # Wire up KPI cards
    ws["B17"] = f"=F21"
    ws["B17"].number_format = fmt_int
    ws["B17"].font      = Font(name="Calibri", bold=True, size=13, color=NAVY)
    ws["B17"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B17"].fill      = solid(BLUE_LIGHT)

    ws["D17"] = f"=F{cval}"
    ws["D17"].number_format = fmt_int
    ws["D17"].font      = Font(name="Calibri", bold=True, size=13, color=NAVY)
    ws["D17"].alignment = Alignment(horizontal="center", vertical="center")
    ws["D17"].fill      = solid(GOLD_LIGHT)

    ws["F17"] = f"=F{cval}-F21-F{net_row}"
    ws["F17"].number_format = f'{sym}#,##0;({sym}#,##0);-'
    ws["F17"].font      = Font(name="Calibri", bold=True, size=13, color=NAVY)
    ws["F17"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F17"].fill      = solid(GREY_LIGHT)

    ws["G17"] = f"=IF(F21=0,0,(F{cval}-F21-F{net_row})/F21)"
    ws["G17"].number_format = pct_fmt
    ws["G17"].font      = Font(name="Calibri", bold=True, size=13, color=GREEN)
    ws["G17"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G17"].fill      = solid(GREEN_BG)

    # Footer
    ws.row_dimensions[pr + 1].height = 5
    ws.row_dimensions[pr + 2].height = 18
    ws.merge_cells(f"B{pr+2}:G{pr+2}")
    cell(ws, pr + 2, 2,
         "This statement is for informational purposes only and does not constitute an offer or solicitation. "
         "Past performance is not indicative of future results.",
         italic=True, size=8, color=TEXT_MED, bg=GREY_LIGHT, align="center", wrap=True)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
    ws.page_margins.left  = 0.5
    ws.page_margins.right = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def send_email(investor: dict, excel_bytes: bytes) -> tuple:
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    from_name = os.environ.get("FROM_NAME", "StatementIQ")

    if not smtp_user or not smtp_pass:
        return False, "SMTP credentials not configured"

    to_email = investor.get("email", "")
    if not to_email:
        return False, "No investor email provided"

    lp_name   = investor.get("lp_name", "Investor")
    fund_name = investor.get("fund_name", "the Fund")
    period_t  = investor.get("period_to", "")
    curr_code = investor.get("currency", "USD")

    msg            = MIMEMultipart()
    msg["From"]    = f"{from_name} <{smtp_user}>"
    msg["To"]      = to_email
    msg["Subject"] = f"Capital Account Statement — {fund_name} — Period Ending {period_t}"

    body = f"""Dear {lp_name},

Please find attached your Limited Partner Capital Account Statement for {fund_name}, covering the period ending {period_t}.

The statement includes:
  • Opening capital account balance
  • Capital calls and distributions during the period
  • Closing capital account value
  • Net gain / (loss) for the period

Currency: {curr_code}

This statement is for informational purposes only. Please contact us if you have any questions or require clarification.

Best regards,
{fund_name}

---
Confidential — For the named LP only. Not for redistribution.
Past performance is not indicative of future results.
"""
    msg.attach(MIMEText(body, "plain"))

    filename = (f"LP_Statement_{investor.get('lp_name','LP').replace(' ','_')}"
                f"_{fund_name.replace(' ','_')}_{period_t}.xlsx")
    part = MIMEBase("application", "octet-stream")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.ehlo(); s.starttls(); s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, to_email, msg.as_string())
        return True, "Email sent"
    except Exception as e:
        return False, str(e)


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_POST(self):
        try:
            length  = int(self.headers.get("Content-Length", 0))
            payload = self.rfile.read(length)
            data    = json.loads(payload)
        except Exception as e:
            self._json(400, {"error": f"Invalid JSON: {e}"}); return

        investor = data.get("investor", {})
        if not investor:
            self._json(400, {"error": "Missing 'investor' in payload"}); return
        if not OPENPYXL_AVAILABLE:
            self._json(500, {"error": "openpyxl not installed"}); return

        try:
            excel_bytes = generate_excel(investor)
        except Exception as e:
            self._json(500, {"error": f"Generation failed: {e}"}); return

        email_sent, email_msg = send_email(investor, excel_bytes)
        filename = (f"LP_Statement_{investor.get('lp_name','LP').replace(' ','_')}"
                    f"_{investor.get('period_to','')}.xlsx")

        self._json(200, {
            "success":       True,
            "filename":      filename,
            "file_b64":      base64.b64encode(excel_bytes).decode("utf-8"),
            "email_sent":    email_sent,
            "email_message": email_msg,
        })

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json(self, status, body):
        enc = json.dumps(body).encode("utf-8")
        self.send_response(status); self._cors()
        self.send_header("Content-Type",   "application/json")
        self.send_header("Content-Length", str(len(enc)))
        self.end_headers(); self.wfile.write(enc)

    def log_message(self, *args): pass
